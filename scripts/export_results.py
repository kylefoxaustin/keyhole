"""
Export all bake-off JSONs to a single XLSX workbook with one sheet per
bake-off. Companion to REPRODUCE.md — gives a SWE a pivotable, shareable
copy of every number in the deck.

Output: data/output/keyhole_results.xlsx

Sheets (13, all under the 31-char XLSX name limit):
  Index                     — cover / description
  Platform                  — host hardware specs (live-probed)
  Mask bake-off             — mobilesam / es-tiny / es-small / yolo-seg IoU + FPS
  FP8 quantization          — ES-Small & YOLO-seg FP8 vs BF16
  INT8 + SmoothQuant        — same, plus SmoothQuant attempt (blocked)
  Hybrid V2 CLIP quant      — CLIP FP8/INT8 on the new architecture
  CLIP keyframe debounce    — N-sweep with FPS + stability
  YOLO conv quant torchao   — 1x1 Conv -> Linear INT8/FP8 attempt
  TRT YOLO                  — TensorRT FP16/INT8/FP8 engine measurements
  TRT CLIP                  — TensorRT FP16/FP8 on CLIP visual tower
  LLM Qwen3 5090            — synthetic prefill + decode + RAG
  LLM NPU tier actuals      — vendor Low/Mid/High benchmarks
  Concurrency               — YOLO batching + deployment scenarios
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

BAKEOFF_DIR = REPO_ROOT / "data" / "output" / "bakeoff"
OUT_PATH = REPO_ROOT / "data" / "output" / "keyhole_results.xlsx"


# ───────────────────────── Helpers ─────────────────────────

def _load(name: str) -> dict | None:
    p = BAKEOFF_DIR / name
    return json.loads(p.read_text()) if p.exists() else None


def _platform_specs() -> list[dict]:
    """Mirror what slide_platform_specs reads — live-probed."""
    from scripts.build_deck import gather_platform_specs
    s = gather_platform_specs()
    rows = [
        ["System",          s["system"]],
        ["Motherboard",     s["motherboard"]],
        ["OS",              s["os"]],
        ["Kernel",          s["kernel"]],
        ["CPU model",       s["cpu_model"]],
        ["CPU cores",       s["cpu_cores"]],
        ["CPU threads",     s["cpu_threads"]],
        ["CPU max MHz",     s["cpu_max_mhz"]],
        ["CPU L3",          s["cpu_l3"]],
        ["RAM (GB)",        s["ram_gb"]],
    ] + [["Storage", ln] for ln in s["storage"]] + [
        ["GPU",             s["gpu"].get("name", "")],
        ["GPU compute cap", s["gpu"].get("compute_cap", "")],
        ["GPU SMs",         s["gpu"].get("sm_count", "")],
        ["GPU VRAM (MiB)",  s["gpu"].get("vram_mib", "")],
        ["GPU TDP (W)",     s["gpu"].get("tdp_w", "")],
        ["GPU SM clock (MHz)",    s["gpu"].get("sm_clock_mhz", "")],
        ["GPU Mem clock (MHz)",   s["gpu"].get("mem_clock_mhz", "")],
        ["GPU L2 cache (MB)",     s["gpu"].get("l2_cache_mb", "")],
        ["CUDA runtime",          s["gpu"].get("cuda_runtime", "")],
        ["cuDNN",                 s["gpu"].get("cudnn", "")],
        ["Driver",                s["gpu"].get("driver", "")],
    ]
    return [{"Property": k, "Value": v} for k, v in rows]


# ───────────────────────── Per-bake-off extractors ─────────────────────────

def sheet_index() -> pd.DataFrame:
    return pd.DataFrame([
        {"Sheet": "Index",                    "Description": "This sheet"},
        {"Sheet": "Platform",                 "Description": "Host hardware on which all 5090 numbers were measured"},
        {"Sheet": "Mask bake-off",            "Description": "MobileSAM / EfficientSAM-Tiny / -Small / YOLO-seg — IoU, params, VRAM, 5090 FPS, edge-projected FPS"},
        {"Sheet": "FP8 quantization",         "Description": "torchao FP8 on ES-Small & YOLO-seg — IoU delta, edge FPS gain"},
        {"Sheet": "INT8 + SmoothQuant",       "Description": "Plain INT8 & SmoothQuant attempt on same contestants"},
        {"Sheet": "Hybrid V2 CLIP quant",     "Description": "CLIP FP8/INT8 on Hybrid V2 (YOLO-seg + CLIP) — concept tag stability"},
        {"Sheet": "CLIP keyframe debounce",   "Description": "Run CLIP every Nth frame — stability vs effective edge FPS"},
        {"Sheet": "YOLO conv quant torchao",  "Description": "swap_conv2d_1x1_to_linear + INT8 on YOLO-seg; FP8 blocked"},
        {"Sheet": "EfficientSAM3 SAM3 Lite",  "Description": "Apr 2026 community SAM 3 Lite (ES-EV-S): 5090 ms + NPU Mid projection + IoU vs SAM 3"},
        {"Sheet": "YOLOE-26 one-model",       "Description": "Jan 2026 Ultralytics YOLOE-26 open-vocab: text-prompted vs prompt-free, 5090 ms + NPU Mid FPS + box recall"},
        {"Sheet": "TRT YOLO",                 "Description": "TensorRT FP16/INT8/FP8 on YOLO-seg — full Conv unblock on Blackwell"},
        {"Sheet": "TRT CLIP",                 "Description": "TensorRT FP16/FP8 on CLIP ViT-B-32 visual tower"},
        {"Sheet": "LLM Qwen3 5090",           "Description": "Qwen3-30B-A3B Q4_K_M/Q5_K_M/Q8_0 — prefill sweep, decode sweep, RAG"},
        {"Sheet": "LLM NPU tier actuals",     "Description": "Vendor NPU Low/Mid/High actuals; per-quant derived decode & RAG times"},
        {"Sheet": "Concurrency",              "Description": "YOLO batched inference B=1..16 + multi-stream deployment scenarios"},
    ])


def sheet_platform() -> pd.DataFrame:
    return pd.DataFrame(_platform_specs())


def sheet_mask_bakeoff() -> pd.DataFrame | None:
    edge = _load("edge_projection.json")
    if not edge:
        return None
    rows = []
    clips = ["720p", "1080p", "4K"]
    for res in clips:
        for name, data in edge["projections"].get(res, {}).items():
            stem = {"720p": "720p_EW_clip",
                    "1080p": "embedded_world_clip_1080p",
                    "4K": "embedded_world_clip"}[res]
            summary_path = BAKEOFF_DIR / stem / "summary.json"
            mean_iou = None
            if summary_path.exists():
                ssum = json.loads(summary_path.read_text())
                c = ssum.get("contestants", {}).get(name, {})
                mean_iou = c.get("mean_iou")
            rows.append({
                "Model": name,
                "Resolution": res,
                "Params (M)": round(data.get("params_m", 0), 2),
                "Mean IoU vs SAM 3": round(mean_iou, 4) if mean_iou else None,
                "GFLOPs/frame": data.get("gflops_per_frame"),
                "5090 ms/frame": round(data.get("measured_frame_ms_5090", 0), 2),
                "5090 FPS": round(data.get("measured_fps_5090", 0), 1),
                "Edge ms/frame (projected)": round(data.get("projected_frame_ms_edge", 0), 1),
                "Edge FPS (projected)": round(data.get("projected_fps_edge", 0), 2),
                "Bottleneck": data.get("bottleneck"),
                "Fits in memory": data.get("fits_in_memory"),
                "Memory headroom (MB)": round(data.get("memory_headroom_mb", 0), 1),
            })
    return pd.DataFrame(rows)


def sheet_fp8() -> pd.DataFrame | None:
    d = _load("fp8_edge_projection.json")
    if not d:
        return None
    rows = []
    for res, by_model in d.get("fp8_projections", {}).items():
        for name, p in by_model.items():
            rows.append({
                "Model": name, "Resolution": res,
                "FP8 applied": p.get("fp8_actually_applied"),
                "Linears swapped": p.get("n_fp8_weights_swapped"),
                "IoU BF16": round(p.get("mean_iou_bf16", 0), 4),
                "IoU FP8":  round(p.get("mean_iou_fp8", 0), 4),
                "Delta IoU": round(p.get("iou_delta", 0), 4),
                "5090 ms BF16": round(p.get("measured_frame_ms_5090_bf16", 0), 2),
                "5090 ms FP8":  round(p.get("measured_frame_ms_5090_fp8", 0), 2),
                "Edge ms BF16 (proj)": round(p.get("projected_ms_edge_bf16", 0), 1),
                "Edge ms FP8 (proj)":  round(p.get("projected_ms_edge_fp8", 0), 1),
                "Edge FPS FP8 (proj)": round(p.get("projected_fps_edge_fp8", 0), 2),
                "Compute-limited ms": round(p.get("compute_limited_ms", 0), 2),
                "BW-limited ms BF16": round(p.get("bandwidth_limited_ms_bf16", 0), 2),
                "BW-limited ms FP8":  round(p.get("bandwidth_limited_ms_fp8", 0), 2),
            })
    return pd.DataFrame(rows)


def sheet_smoothquant() -> pd.DataFrame | None:
    d = _load("smoothquant_edge_projection.json")
    if not d:
        return None
    rows = []
    for res, by_model in d.get("projections", {}).items():
        for name, by_recipe in by_model.items():
            for recipe, p in by_recipe.items():
                rows.append({
                    "Model": name, "Resolution": res, "Recipe": recipe,
                    "Applied": p.get("recipe_applied"),
                    "Linears quantized": p.get("n_quantized"),
                    "IoU BF16": round(p.get("mean_iou_bf16", 0), 4),
                    "IoU recipe": round(p.get("mean_iou_recipe", 0), 4),
                    "Delta IoU": round(p.get("iou_delta", 0), 4),
                    "5090 ms BF16": round(p.get("measured_frame_ms_5090_bf16", 0), 2),
                    "5090 ms recipe": round(p.get("measured_frame_ms_5090_recipe", 0), 2),
                    "Edge ms BF16 (proj)": round(p.get("projected_ms_edge_bf16", 0), 1),
                    "Edge ms recipe (proj)": round(p.get("projected_ms_edge_recipe", 0), 1),
                    "Edge FPS recipe (proj)": round(p.get("projected_fps_edge_recipe", 0), 2),
                })
    return pd.DataFrame(rows)


def sheet_hybrid_v2_clip() -> pd.DataFrame | None:
    d = _load("hybrid_v2_edge_projection.json")
    if not d:
        return None
    rows = []
    for res, by_recipe in d.get("projections", {}).items():
        for recipe, p in by_recipe.items():
            rows.append({
                "Resolution": res, "Recipe": recipe,
                "CLIP Linears quantized": p.get("n_quantized"),
                "CLIP total Linears": p.get("n_linear"),
                "Top-1 agreement vs BF16": round(p.get("top1_agreement", 0), 4),
                "Top-3 Jaccard vs BF16":   round(p.get("top3_jaccard", 0), 4),
                "5090 YOLO ms": round(p.get("mean_yolo_ms_5090", 0), 2),
                "5090 CLIP ms": round(p.get("mean_clip_ms_5090", 0), 2),
                "5090 total ms": round(p.get("mean_total_ms_5090", 0), 2),
                "Edge YOLO ms (proj)": round(p.get("projected_yolo_ms_edge", 0), 1),
                "Edge CLIP ms (proj)": round(p.get("projected_clip_ms_edge", 0), 1),
                "Edge total ms (proj)": round(p.get("projected_total_ms_edge", 0), 1),
                "Edge FPS (proj)": round(p.get("projected_fps_edge", 0), 2),
            })
    return pd.DataFrame(rows)


def sheet_yoloe26() -> pd.DataFrame | None:
    d = _load("yoloe26_summary.json")
    if not d:
        return None
    bw_ratio = (1792.0 * 0.85) / (134.4 * 0.80)
    rows = []
    for tag, v in d.get("variants", {}).items():
        if "error" in v:
            continue
        for res, r in v.get("by_resolution", {}).items():
            ms_5090 = r["per_frame_ms_5090"]["p50"]
            ms_mid = ms_5090 * bw_ratio
            rows.append({
                "Variant":           tag,
                "Weights":           v["weights"],
                "Prompt free":       v["prompt_free"],
                "Resolution":        res,
                "Params (M)":        round(v["params_m"], 2),
                "VRAM MB (5090)":    round(v["peak_vram_mb_5090"], 0),
                "5090 mean ms":      round(r["per_frame_ms_5090"]["mean"], 2),
                "5090 p50 ms":       round(r["per_frame_ms_5090"]["p50"], 2),
                "5090 p95 ms":       round(r["per_frame_ms_5090"]["p95"], 2),
                "NPU Mid ms (BW-scaled)": round(ms_mid, 1),
                "NPU Mid FPS":       round(1000.0 / ms_mid, 2) if ms_mid > 0 else 0.0,
                "N YOLOE boxes":     r["n_boxes_yoloe"],
                "N ref YOLO11x boxes": r["n_boxes_reference_yolo11x"],
                "N matched IoU≥0.5": r["n_matched_boxes_iou_ge_0.5"],
                "Box recall":        round(r["box_recall_vs_yolo11x"], 3),
            })
    return pd.DataFrame(rows)


def sheet_efficientsam3() -> pd.DataFrame | None:
    d = _load("efficientsam3_summary.json")
    if not d:
        return None
    bw_ratio = (1792.0 * 0.85) / (134.4 * 0.80)   # 5090 eff / NPU Mid eff ≈ 14.17
    rows = []
    for res, r in d.get("by_resolution", {}).items():
        ms_5090 = r["per_frame_ms_5090"]["p50"]
        ms_mid = ms_5090 * bw_ratio
        rows.append({
            "Resolution":        res,
            "Clip":              r["clip"],
            "N frames timed":    r["n_frames_timed"],
            "N boxes (total)":   r["n_boxes_total"],
            "5090 mean ms":      round(r["per_frame_ms_5090"]["mean"], 1),
            "5090 p50 ms":       round(r["per_frame_ms_5090"]["p50"], 1),
            "5090 p95 ms":       round(r["per_frame_ms_5090"]["p95"], 1),
            "NPU Mid ms (BW-scaled)": round(ms_mid, 0),
            "NPU Mid FPS (proj)": round(1000.0 / ms_mid, 2) if ms_mid > 0 else 0.0,
            "Mean IoU vs SAM 3": round(r["iou_vs_sam3"]["mean"], 3),
            "Median IoU vs SAM 3": round(r["iou_vs_sam3"]["median"], 3),
            "IoU sample count":  r["iou_vs_sam3"]["n"],
        })
    return pd.DataFrame(rows)


def sheet_keyframe_debounce() -> pd.DataFrame | None:
    d = _load("keyframe_debounce_summary.json")
    if not d:
        return None
    rows = []
    for res, data in d.get("per_resolution", {}).items():
        for r in data.get("rows", []):
            rows.append({
                "Resolution": res,
                "N (native frames)": r["N_native_frames"],
                "Keyframe interval (sec)": r["keyframe_interval_sec"],
                "Gap sampled used": r.get("gap_sampled_used"),
                "Top-1 stability": round(r["stability_top1"], 3),
                "Top-3 Jaccard stability": round(r.get("stability_top3_jaccard", 0), 3),
                "Pairs in measurement": r.get("n_pairs"),
                "Eff 5090 ms/frame": round(r.get("eff_5090_ms_per_frame", 0), 2),
                "Eff edge ms/frame (proj)": round(r["eff_edge_ms_per_frame"], 1),
                "Eff edge FPS (proj)": round(r["eff_edge_fps"], 2),
            })
    return pd.DataFrame(rows)


def sheet_yolo_conv_quant() -> pd.DataFrame | None:
    d = _load("yolo_conv_quant_edge_projection.json")
    if not d:
        return None
    rows = []
    for res, by_recipe in d.get("projections", {}).items():
        for recipe, p in by_recipe.items():
            if "error" in p:
                rows.append({"Resolution": res, "Recipe": recipe,
                             "Error": p["error"],
                             "1x1 swapped": None, "Quantized": None})
                continue
            rows.append({
                "Resolution": res, "Recipe": recipe,
                "Total Conv2d": p.get("n_conv2d"),
                "1x1 Convs": p.get("n_conv2d_1x1"),
                "1x1 swapped to Linear": p.get("n_swapped_linears"),
                "Quantized": p.get("n_quantized"),
                "Frac conv weights quantized": round(p.get("frac_conv_weights_quantized", 0), 4),
                "Box recall vs BF16": round(p.get("box_recall", 0), 4),
                "Matched IoU": round(p.get("mean_matched_iou", 0), 4),
                "5090 ms/frame": round(p.get("mean_frame_ms_5090", 0), 2),
                "Edge ms BF16 (proj)": round(p.get("projected_ms_edge_bf16", 0), 1),
                "Edge ms recipe (proj)": round(p.get("projected_ms_edge_recipe", 0), 1),
                "Edge FPS recipe (proj)": round(p.get("projected_fps_edge_recipe", 0), 2),
            })
    return pd.DataFrame(rows)


def sheet_trt_yolo() -> pd.DataFrame | None:
    d = _load("trt_yolo_edge_projection.json")
    if not d:
        return None
    rows = []
    for res, by_recipe in d.get("projections", {}).items():
        for recipe, p in by_recipe.items():
            if "error" in p:
                rows.append({"Resolution": res, "Recipe": recipe,
                             "Error": p["error"]})
                continue
            rows.append({
                "Resolution": res, "Recipe": recipe.upper(),
                "5090 ms/frame": round(p.get("mean_frame_ms_5090", 0), 2),
                "Box recall vs FP16": round(p.get("box_recall", 1.0), 4),
                "Matched IoU": round(p.get("mean_matched_iou", 1.0), 4),
                "Edge ms/frame (proj)": round(p.get("projected_ms_edge", 0), 1),
                "Edge FPS (proj)": round(p.get("projected_fps_edge", 0), 2),
                "BW multiplier": p.get("bandwidth_multiplier"),
                "Compute-limited ms": round(p.get("compute_limited_ms", 0), 2),
                "BW-limited ms": round(p.get("bandwidth_limited_ms", 0), 2),
            })
    return pd.DataFrame(rows)


def sheet_trt_clip() -> pd.DataFrame | None:
    d = _load("trt_clip_edge_projection.json")
    if not d:
        return None
    rows = []
    for res, by_recipe in d.get("projections", {}).items():
        for recipe, p in by_recipe.items():
            if "error" in p:
                rows.append({"Resolution": res, "Recipe": recipe,
                             "Error": p["error"]})
                continue
            rows.append({
                "Resolution": res, "Recipe": recipe,
                "5090 ms/frame": round(p.get("mean_frame_ms_5090", 0), 2),
                "Top-1 agreement vs BF16": round(p.get("top1_agreement", 1.0), 4),
                "Edge CLIP ms (proj)": round(p.get("projected_clip_ms_edge", 0), 1),
                "Edge CLIP-only FPS (proj)": round(p.get("projected_fps_edge_clip_only", 0), 2),
                "BW multiplier": p.get("bandwidth_multiplier"),
            })
    return pd.DataFrame(rows)


def sheet_llm_qwen3_5090() -> pd.DataFrame | None:
    d = _load("llm_summary.json")
    if not d:
        return None
    rows = []
    for quant, data in d.items():
        if "error" in data:
            rows.append({"Quant": quant, "Error": data["error"]})
            continue
        base = {
            "Quant": quant,
            "GGUF file": data.get("gguf_file"),
            "GGUF size (GB)": round(data.get("gguf_size_gb", 0), 2),
            "n_gpu_layers": data.get("n_gpu_layers"),
            "n_ctx": data.get("n_ctx"),
            "Offload note": data.get("offload_note"),
            "Load (sec)": round(data.get("load_sec", 0), 2),
            "Peak VRAM (MB)": round(data.get("peak_vram_mb_during_bench", 0), 0),
        }
        for pr in data.get("prefill_sweep", []):
            rows.append({**base,
                         "Phase": "prefill",
                         "N tokens": pr["n_prompt"],
                         "ms": round(pr["prefill_ms"], 2),
                         "tok/s": round(pr["prefill_tok_s"], 1)})
        for pr in data.get("decode_sweep", []):
            rows.append({**base,
                         "Phase": "decode",
                         "N tokens": pr["n_decode"],
                         "ms": round(pr["decode_ms"], 2),
                         "tok/s": round(pr["decode_tok_s"], 1)})
        rag = data.get("rag", {})
        if rag:
            rows.append({**base, "Phase": "RAG prefill",
                         "N tokens": rag["rag_prompt_len"],
                         "ms": round(rag["prefill_ms"], 2),
                         "tok/s": round(rag["prefill_tok_s"], 1)})
            rows.append({**base, "Phase": "RAG decode",
                         "N tokens": rag["rag_response_len"],
                         "ms": round(rag["decode_ms"], 2),
                         "tok/s": round(rag["decode_tok_s"], 1)})
            rows.append({**base, "Phase": "RAG total",
                         "N tokens": rag["rag_prompt_len"] + rag["rag_response_len"],
                         "ms": round(rag["total_ms"], 2),
                         "tok/s": None})
    return pd.DataFrame(rows)


def sheet_llm_npu_tiers() -> pd.DataFrame | None:
    d = _load("llm_edge_projection.json")
    if not d or "tier_projections" not in d:
        return None
    rows = []
    for tier, tp in d["tier_projections"].items():
        for quant, qd in tp["per_quant"].items():
            rows.append({
                "NPU tier": tier,
                "Memory bus": tp["bus"],
                "Quant": quant,
                "TTFT 1K prompt (sec)": tp["reference_ttft_1k_sec"],
                "Prefill 1K (tok/s)": round(qd.get("prefill_1k_tok_s", 0), 1),
                "Decode (tok/s)": round(qd["decode_tok_s"], 2),
                "Short answer 200 tok (ms)": round(qd["short_answer_ms"], 0),
                "RAG 8K+2K prefill (ms)": round(qd["rag_prefill_ms"], 0),
                "RAG 8K+2K decode (ms)": round(qd["rag_decode_ms"], 0),
                "RAG 8K+2K total (sec)": round(qd["rag_total_sec"], 1),
            })
    return pd.DataFrame(rows)


def sheet_concurrency() -> pd.DataFrame | None:
    d = _load("concurrency_edge_projection.json")
    if not d:
        return None
    rows = []
    # Block 1 — raw batched 5090
    for r in d.get("batches_5090", []):
        rows.append({
            "Section": "5090 raw batched",
            "Batch": r["batch"],
            "Mean ms": round(r["mean_ms"], 2),
            "P95 ms": round(r.get("p95_ms", 0), 2),
            "ms per stream": round(r["per_stream_ms"], 2),
            "Streams per sec": round(r["throughput_streams_per_sec"], 1),
            "Scenario": None,
            "N streams": None,
            "YOLO batch": None,
            "Batch ms edge": None,
            "FPS per stream": None,
            "Total system FPS": None,
        })
    # Block 2 — edge-projected batched
    for r in d.get("batches_edge", []):
        rows.append({
            "Section": "Edge projected",
            "Batch": r["batch"],
            "Mean ms": round(r["mean_ms_edge"], 1),
            "P95 ms": None,
            "ms per stream": round(r["per_stream_ms_edge"], 1),
            "Streams per sec": round(r["streams_per_sec_edge"], 1),
            "Scenario": None,
            "N streams": None,
            "YOLO batch": None,
            "Batch ms edge": None,
            "FPS per stream": None,
            "Total system FPS": None,
        })
    # Block 3 — deployment scenarios
    for s in d.get("scenarios_edge", []):
        rows.append({
            "Section": "Deployment scenario",
            "Batch": None, "Mean ms": None, "P95 ms": None,
            "ms per stream": None, "Streams per sec": None,
            "Scenario": s["label"],
            "N streams": s["n_streams"],
            "YOLO batch": s["yolo_batch"],
            "Batch ms edge": round(s["batch_ms_edge"], 1),
            "FPS per stream": round(s["fps_per_stream"], 1),
            "Total system FPS": round(s["total_system_fps"], 1),
        })
    return pd.DataFrame(rows)


# ───────────────────────── Main ─────────────────────────

SHEETS = [
    ("Index",                    sheet_index),
    ("Platform",                 sheet_platform),
    ("Mask bake-off",            sheet_mask_bakeoff),
    ("FP8 quantization",         sheet_fp8),
    ("INT8 + SmoothQuant",       sheet_smoothquant),
    ("Hybrid V2 CLIP quant",     sheet_hybrid_v2_clip),
    ("CLIP keyframe debounce",   sheet_keyframe_debounce),
    ("YOLO conv quant torchao",  sheet_yolo_conv_quant),
    ("EfficientSAM3 SAM3 Lite",  sheet_efficientsam3),
    ("YOLOE-26 one-model",       sheet_yoloe26),
    ("TRT YOLO",                 sheet_trt_yolo),
    ("TRT CLIP",                 sheet_trt_clip),
    ("LLM Qwen3 5090",           sheet_llm_qwen3_5090),
    ("LLM NPU tier actuals",     sheet_llm_npu_tiers),
    ("Concurrency",              sheet_concurrency),
]


def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    skipped: list[str] = []
    written: list[tuple[str, int]] = []

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for name, fn in SHEETS:
            df = fn()
            if df is None or df.empty:
                skipped.append(name)
                continue
            df.to_excel(writer, sheet_name=name[:31], index=False)
            written.append((name, len(df)))

    # Auto-widen columns for readability
    from openpyxl import load_workbook
    wb = load_workbook(OUT_PATH)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    wb.save(OUT_PATH)

    print(f"\nWrote {OUT_PATH}")
    print(f"\n{'Sheet':28s}  {'Rows':>6s}")
    print("-" * 40)
    for n, r in written:
        print(f"{n:28s}  {r:>6d}")
    if skipped:
        print(f"\nSkipped (JSON not present): {', '.join(skipped)}")


if __name__ == "__main__":
    main()
